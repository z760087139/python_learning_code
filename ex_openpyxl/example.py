import openpyxl
# load from excel
wb = openpyxl.load_workbook('tmp.xlsx')
# wb = openpyxl.load_workbook(..,read_only=True|write_only=True)
# useful wb function
# wb.active , wb.close() , wb.sheetnames , wb['Sheet1']

ws = wb.active
# useful ws function
# ws[1] , ws['A'] , ws.rows , ws.cols , ws.values

# create excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'new_title'
new_ws = wb.create_sheet('create_sheet')
new_ws.append([[1,2,3],[4,5,6]]) # can be list tuple range dict

# vaildation data
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)
dv.error ='Your entry is not in the list'
dv.errorTitle = 'Invalid Entry'
# Optionally set a custom prompt message
dv.prompt = 'Please select from the list'
dv.promptTitle = 'List Selection'
# Add the data-validation object to the worksheet
ws.add_data_validation(dv)
dv.add(c2)

# merge cells / unmerge cells
ws.merge_cells('A2:D2')
ws.unmerge_cells('A2:D2')

def append(self, iterable):
        """Appends a group of values at the bottom of the current sheet.

        * If it's a list: all values are added in order, starting from the first column
        * If it's a dict: values are assigned to the columns indicated by the keys (numbers or letters)

        :param iterable: list, range or generator, or dict containing values to append
        :type iterable: list|tuple|range|generator or dict

        Usage:

        * append(['This is A1', 'This is B1', 'This is C1'])
        * **or** append({'A' : 'This is A1', 'C' : 'This is C1'})
        * **or** append({1 : 'This is A1', 3 : 'This is C1'})

        :raise: TypeError when iterable is neither a list/tuple nor a dict

        """
    row_idx = self._current_row + 1

    if (isinstance(iterable, (list, tuple, range))
        or isgenerator(iterable)):
        for col_idx, content in enumerate(iterable, 1):
            if isinstance(content, Cell):
                # compatible with write-only mode
                cell = content
                if cell.parent and cell.parent != self:
                    raise ValueError("Cells cannot be copied from other worksheets")
                cell.parent = self
                cell.col_idx = col_idx
                cell.row = row_idx
            else:
                cell = Cell(self, row=row_idx, col_idx=col_idx, value=content)
            self._cells[(row_idx, col_idx)] = cell

    elif isinstance(iterable, dict):
        for col_idx, content in iterable.items():
            if isinstance(col_idx, basestring):
                col_idx = column_index_from_string(col_idx)
            cell = Cell(self, row=row_idx, col_idx=col_idx, value=content)
            self._cells[(row_idx, col_idx)] = cell

    else:
        self._invalid_row(iterable)

    self._current_row = row_idx